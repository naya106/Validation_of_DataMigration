import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers, Font
from openpyxl.styles.borders import Border, Side
import Configuration as config
from GetResult_MySql import ConnectMySqlSrc, ConnectMySqlTgt


def writeToFile(filepath):
    file = "전환검증결과(template)_V1.0.xlsx"

    wb = load_workbook(os.path.join(config.base_dir, file))
    ws = wb.active
    ws.title = '전환결과'

    font = Font(name='맑은 고딕', size=10)
    alignment = Alignment(horizontal='center', vertical='center')

    number_format = numbers.builtin_format_code(3)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws.merge_cells('E2:F2')
    ws.merge_cells('E3:F3')
    ws.merge_cells('C5:D5')
    ws.merge_cells('E5:F5')
    ws.merge_cells('C6:F6')

    ws['E2'].alignment = alignment
    ws['E3'].alignment = alignment
    ws['C5'].alignment = Alignment('general', 'center', wrap_text=True)
    ws['E5'].alignment = Alignment('general', 'center', wrap_text=True)
    ws['C6'].alignment = alignment

    ws['C2'] = config.module
    ws['C2'].alignment = alignment
    ws['C3'] = config.srcTableNm
    ws['C3'].alignment = alignment
    ws['C5'] = config.srcSql
    ws['E5'] = config.tgtSql
    ws['C6'] = config.migIssue
    ws['E2'] = config.migDate
    ws['E3'] = config.tgtTableNm
    ws['C10'] = '=SUM(C11:C100)'
    ws['C10'].number_format = number_format
    ws['C10'].alignment = alignment
    ws['E10'] = '=SUM(E11:E100)'
    ws['E10'].number_format = number_format
    ws['E10'].alignment = alignment
    ws['F10'] = '=IF(AND(C10-E10=0), "O", "X")'
    ws['F10'].alignment = alignment

    connSrc = ConnectMySqlSrc()
    connTgt = ConnectMySqlTgt()

    src = list(connSrc.fetchRows())
    tgt = list(connTgt.fetchRows())

    start_row = 11
    for row, i in enumerate(src, start_row):
        cell_1 = ws.cell(row=row, column=2, value=i[0])
        cell_1.alignment = alignment
        cell_1.font = font
        cell_2 = ws.cell(row=row, column=3, value=i[1])
        cell_2.alignment = alignment
        cell_2.font = font

    for row, x in enumerate(src, start_row):
        cell_3 = ws.cell(row=row, column=4, value=x[0])
        cell_3.alignment = alignment
        cell_3.font = font
        cell_4 = ws.cell(row=row, column=5, value=x[1])
        cell_4.alignment = alignment
        cell_4.font = font

    for row, (l1, l2) in enumerate(zip(src, tgt), start_row):
        if l1[1] - l2[1] == 0:
            cell_5 = ws.cell(row=row, column=6, value="O")
            cell_5.alignment = alignment
            cell_5.font = font
        else:
            cell_6 = ws.cell(row=row, column=6, value="X")
            cell_6.alignment = alignment
            cell_6.font = font

    if ws.cell(row=ws.max_row, column=5).value > 0:
        for row in ws.iter_rows(min_row=8, min_col=2, max_row=ws.max_row, max_col=6):
            for cell in row:
                cell.border = thin_border

    wb.save(filepath)
    wb.close()


if __name__ == "__main__":
    export = config.exportPath
    fileNm = "전환검증결과(" + config.tgtTableNm + ")_V1.0.xlsx"
    writeToFile(os.path.join(export, fileNm))